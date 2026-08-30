"""Bikin docs/data-penduduk.xlsx — data CONTOH untuk menguji sistem.

Isinya karangan: 100 keluarga utuh (bapak/ibu/anak) yang berbagi alamat dan
marga, umur & pendidikan & pekerjaan dibuat saling masuk akal. Bukan warga
sungguhan. NIK & No. KK tidak dibangkitkan — sistem tidak menyimpannya.

Bukan bagian aplikasi; alat sekali jalan. Bentuk filenya sama persis dengan
yang diisi pengurus, jadi jalur impornya juga sama.

Pakai (dari backend/):
    .venv/bin/python -m tools.buat_data_contoh
"""

import random
from datetime import date, timedelta

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from app.data.impor_excel import BARIS_HEADER, KOLOM, NAMA_SHEET

random.seed(20260819)

DESA, KECAMATAN, KABUPATEN, PROVINSI, KODE_POS = (
    "Sukamaju", "Cibiru", "Bandung", "Jawa Barat", "40615",
)
# RW 019 -> RT 001,002 | RW 020 -> RT 003,004 | RW 021 -> RT 005,006
RW_RT = {"019": ["001", "002"], "020": ["003", "004"], "021": ["005", "006"]}
JUMLAH_KK = 100

NAMA_DEPAN_L = [
    "Agus", "Bambang", "Budi", "Dedi", "Eko", "Fajar", "Gunawan", "Hendra",
    "Irfan", "Joko", "Kurnia", "Lukman", "Maman", "Nanang", "Oman", "Purnomo",
    "Rahmat", "Slamet", "Taufik", "Usman", "Wahyu", "Yusuf", "Zainal", "Asep",
    "Dadang", "Endang", "Ferry", "Hadi", "Imam", "Jaka",
]
NAMA_DEPAN_P = [
    "Ani", "Betty", "Citra", "Dewi", "Endah", "Fitri", "Gita", "Hesti",
    "Indah", "Juwita", "Kartika", "Lestari", "Mira", "Nurul", "Oktavia",
    "Puspita", "Ratna", "Siti", "Tuti", "Umi", "Wulan", "Yanti", "Zahra",
    "Ade", "Diah", "Eka", "Feni", "Hani", "Ika", "Jamilah",
]
NAMA_ANAK_L = [
    "Adit", "Bayu", "Cahyo", "Dimas", "Erlangga", "Farhan", "Galih", "Hafiz",
    "Ilham", "Jefri", "Kevin", "Lutfi", "Mahesa", "Naufal", "Okta", "Pandu",
    "Rizky", "Satria", "Teguh", "Umar", "Vino", "Wisnu", "Yoga", "Zidan",
]
NAMA_ANAK_P = [
    "Alya", "Bunga", "Cinta", "Dara", "Elsa", "Fani", "Gina", "Hana",
    "Intan", "Jihan", "Keisha", "Laras", "Maya", "Nadia", "Olivia", "Prita",
    "Rani", "Salma", "Tiara", "Ulfa", "Vera", "Widya", "Yuni", "Zaskia",
]
MARGA = [
    "Pratama", "Wijaya", "Nugraha", "Saputra", "Kurniawan", "Hidayat",
    "Ramadhan", "Setiawan", "Firmansyah", "Maulana", "Santoso", "Gunawan",
    "Permana", "Susanto", "Herlambang", "Prasetyo", "Wibowo", "Sanjaya",
]
JALAN = [
    "Jl. Melati", "Jl. Mawar", "Jl. Anggrek", "Jl. Kenanga", "Jl. Dahlia",
    "Jl. Flamboyan", "Jl. Cempaka", "Jl. Bougenville", "Jl. Teratai",
    "Gg. Masjid", "Gg. Sukamaju", "Jl. Raya Cibiru",
]
KERJA_L = [
    "Petani", "Buruh Harian Lepas", "Wiraswasta", "Karyawan Swasta",
    "Pedagang", "Sopir", "Tukang Bangunan", "Guru", "Pegawai Negeri Sipil",
    "Montir", "Satpam", "Nelayan",
]
KERJA_P = [
    "Ibu Rumah Tangga", "Ibu Rumah Tangga", "Ibu Rumah Tangga", "Pedagang",
    "Guru", "Karyawan Swasta", "Penjahit", "Buruh Harian Lepas", "Bidan",
]
AGAMA_BOBOT = (
    ["ISLAM"] * 90 + ["KRISTEN"] * 5 + ["KATOLIK"] * 3 + ["HINDU"] + ["BUDDHA"]
)
GOLDAR = ["A", "B", "AB", "O", "TIDAK_TAHU"]
HARI_INI = date(2026, 8, 19)


def tanggal_acak(umur_min: int, umur_maks: int) -> date:
    hari = random.randint(umur_min * 365, umur_maks * 365)
    return HARI_INI - timedelta(days=hari)


def umur(lahir: date) -> int:
    return (HARI_INI - lahir).days // 365


def pendidikan_utk(u: int) -> str:
    if u < 6:
        return "TIDAK_SEKOLAH"
    if u < 12:
        return "SD"
    if u < 15:
        return "SMP"
    if u < 18:
        return "SMA"
    return random.choice(
        ["SD"] * 2 + ["SMP"] * 3 + ["SMA"] * 8 + ["D3"] * 2 + ["S1"] * 3 + ["S2"]
    )


def pekerjaan_utk(u: int, perempuan: bool, pendidikan: str) -> str:
    if u < 6:
        return "Belum/Tidak Bekerja"
    if u < 18:
        return "Pelajar/Mahasiswa"
    if u < 23 and pendidikan in {"D3", "S1", "S2"}:
        return "Pelajar/Mahasiswa"
    if u >= 65:
        return "Pensiunan" if random.random() < 0.4 else "Tidak Bekerja"
    return random.choice(KERJA_P if perempuan else KERJA_L)


baris_semua = []
urut_kode = 1

for i in range(JUMLAH_KK):
    rw = list(RW_RT)[i % 3]
    rt = random.choice(RW_RT[rw])
    alamat_jalan = f"{random.choice(JALAN)} No. {random.randint(1, 90)}"
    marga = random.choice(MARGA)
    agama = random.choice(AGAMA_BOBOT)

    ayah_lahir = tanggal_acak(30, 62)
    # Istri seumuran suami (-3..+5 tahun), bukan acak lepas — pasangan yang
    # selisihnya belasan tahun bikin datanya kelihatan karangan.
    umur_ayah = umur(ayah_lahir)
    ibu_lahir = tanggal_acak(max(22, umur_ayah - 5), max(23, umur_ayah + 3))
    anggota = []

    # Kepala keluarga
    u = umur(ayah_lahir)
    pdd = pendidikan_utk(u)
    anggota.append({
        "lahir": ayah_lahir, "perempuan": False,
        "nama": f"{random.choice(NAMA_DEPAN_L)} {marga}",
        "statusPerkawinan": "KAWIN", "pendidikan": pdd,
        "pekerjaan": pekerjaan_utk(u, False, pdd),
        "statusHubunganKeluarga": "KEPALA_KELUARGA",
    })

    # Istri
    u = umur(ibu_lahir)
    pdd = pendidikan_utk(u)
    anggota.append({
        "lahir": ibu_lahir, "perempuan": True,
        "nama": f"{random.choice(NAMA_DEPAN_P)} {marga}",
        "statusPerkawinan": "KAWIN", "pendidikan": pdd,
        "pekerjaan": pekerjaan_utk(u, True, pdd),
        "statusHubunganKeluarga": "ISTRI",
    })

    # 1-3 anak. Umurnya dibatasi dua sisi terhadap umur ibu: anak sulung tidak
    # lahir sebelum ibunya 20, anak bungsu tidak lahir setelah ibunya 45.
    umur_ibu = umur(ibu_lahir)
    maks_umur_anak = max(1, umur_ibu - 20)
    min_umur_anak = max(0, umur_ibu - 45)
    if min_umur_anak >= maks_umur_anak:
        min_umur_anak = max(0, maks_umur_anak - 1)
    for _ in range(random.randint(1, 3)):
        anak_lahir = tanggal_acak(min_umur_anak, maks_umur_anak)
        ua = umur(anak_lahir)
        perempuan = random.random() < 0.5
        pdd = pendidikan_utk(ua)
        anggota.append({
            "lahir": anak_lahir, "perempuan": perempuan,
            "nama": f"{random.choice(NAMA_ANAK_P if perempuan else NAMA_ANAK_L)} {marga}",
            "statusPerkawinan": "BELUM_KAWIN" if ua < 22 else random.choice(
                ["BELUM_KAWIN", "BELUM_KAWIN", "KAWIN"]
            ),
            "pendidikan": pdd, "pekerjaan": pekerjaan_utk(ua, perempuan, pdd),
            "statusHubunganKeluarga": "ANAK",
        })

    for a in anggota:
        baris_semua.append({
            "id": f"W{urut_kode:04d}",
            "nama": a["nama"],
            "jenisKelamin": "PEREMPUAN" if a["perempuan"] else "LAKI_LAKI",
            "tempatLahir": random.choice(
                ["Bandung", "Bandung", "Bandung", "Garut", "Cimahi", "Sumedang", "Tasikmalaya"]
            ),
            "tanggalLahir": a["lahir"].isoformat(),
            "agama": agama,
            "statusPerkawinan": a["statusPerkawinan"],
            "pendidikan": a["pendidikan"],
            "pekerjaan": a["pekerjaan"],
            "golonganDarah": random.choice(GOLDAR),
            "statusHubunganKeluarga": a["statusHubunganKeluarga"],
            "kewarganegaraan": "WNI",
            "jabatan": "WARGA",  # disetel di bawah untuk 10 pemegang jabatan
            "jalan": alamat_jalan, "rt": rt, "rw": rw,
            "desa": DESA, "kecamatan": KECAMATAN, "kabupaten": KABUPATEN,
            "provinsi": PROVINSI, "kodePos": KODE_POS,
        })
        urut_kode += 1

assert baris_semua, "tidak ada baris terbangkitkan"
assert len({b["id"] for b in baris_semua}) == len(baris_semua), "Kode Warga dobel"

# --- tunjuk pemegang jabatan: kepala keluarga tertua di tiap wilayah -------
# Pilihan yang bisa dijelaskan, bukan baris pertama yang kebetulan ketemu.
kepala = sorted(
    (b for b in baris_semua if b["statusHubunganKeluarga"] == "KEPALA_KELUARGA"),
    key=lambda b: b["tanggalLahir"],
)
terpakai: set[str] = set()


def _tunjuk(kandidat, jabatan: str) -> None:
    for b in kandidat:
        if b["id"] not in terpakai:
            terpakai.add(b["id"])
            b["jabatan"] = jabatan
            return
    raise SystemExit(f"tidak ada kandidat untuk {jabatan}")


_tunjuk(kepala, "DUKUH")
for _rw in sorted({b["rw"] for b in baris_semua}):
    _tunjuk([b for b in kepala if b["rw"] == _rw], "RW")
for _rw, _rt in sorted({(b["rw"], b["rt"]) for b in baris_semua}):
    _tunjuk([b for b in kepala if b["rw"] == _rw and b["rt"] == _rt], "RT")

# --- tulis ke xlsx, bentuknya sama persis dengan template kosong ------------

wb = Workbook()
ws = wb.active
ws.title = NAMA_SHEET

ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(KOLOM))
sel = ws.cell(row=1, column=1, value=f"DATA PENDUDUK — DESA {DESA.upper()}")
sel.fill = PatternFill("solid", fgColor="1F4E78")
sel.font = Font(color="FFFFFF", bold=True, size=13)
sel.alignment = Alignment(horizontal="center", vertical="center")
ws.row_dimensions[1].height = 24

TIPIS = Side(style="thin", color="B7B7B7")
BINGKAI = Border(left=TIPIS, right=TIPIS, top=TIPIS, bottom=TIPIS)

for i, (_, label, lebar) in enumerate(KOLOM, start=1):
    c = ws.cell(row=BARIS_HEADER, column=i, value=label)
    c.fill = PatternFill("solid", fgColor="D9E1F2")
    c.font = Font(bold=True)
    c.border = BINGKAI
    c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    ws.column_dimensions[get_column_letter(i)].width = lebar
ws.row_dimensions[BARIS_HEADER].height = 30
ws.freeze_panes = f"A{BARIS_HEADER + 1}"

for r, baris in enumerate(baris_semua, start=BARIS_HEADER + 1):
    for i, (kunci, _, _) in enumerate(KOLOM, start=1):
        c = ws.cell(row=r, column=i, value=baris[kunci])
        c.border = BINGKAI
        if kunci in {"nik", "noKK"}:
            c.alignment = Alignment(horizontal="left")

ws.auto_filter.ref = f"A{BARIS_HEADER}:{get_column_letter(len(KOLOM))}{len(baris_semua) + BARIS_HEADER}"

wb.save("../docs/data-penduduk.xlsx")

jml_kk = len({b["noKK"] for b in baris_semua})
per_kk = [sum(1 for b in baris_semua if b["noKK"] == kk) for kk in {b["noKK"] for b in baris_semua}]
print(f"OK: {len(baris_semua)} jiwa, {jml_kk} KK, anggota per KK {min(per_kk)}-{max(per_kk)}")
