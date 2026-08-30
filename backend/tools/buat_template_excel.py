"""Bikin docs/template-data-penduduk.xlsx — formulir KOSONG untuk diisi pengurus.

Bukan bagian aplikasi; alat sekali jalan. Definisi kolomnya diambil dari
`app.data.impor_excel` supaya bentuk formulir dan pembacanya tidak mungkin
melenceng sendiri-sendiri.

Pakai (dari backend/):
    .venv/bin/python -m tools.buat_template_excel
"""

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

from app.data.impor_excel import BARIS_HEADER, KOLOM, NAMA_SHEET, PILIHAN

KELUARAN = "../docs/template-data-penduduk.xlsx"
BARIS_KOSONG = 300

CONTOH = [
    {
        "id": "W0001",
        "nama": "Contoh Nama Kepala Keluarga", "jenisKelamin": "LAKI_LAKI",
        "tempatLahir": "Bandung", "tanggalLahir": "1990-01-01", "agama": "ISLAM",
        "statusPerkawinan": "KAWIN", "pendidikan": "SMA", "pekerjaan": "Wiraswasta",
        "golonganDarah": "O", "statusHubunganKeluarga": "KEPALA_KELUARGA",
        "kewarganegaraan": "WNI", "jabatan": "DUKUH",
        "jalan": "Jl. Contoh No. 1", "rt": "001",
        "rw": "019", "desa": "Sukamaju", "kecamatan": "Cibiru",
        "kabupaten": "Bandung", "provinsi": "Jawa Barat", "kodePos": "40615",
    },
    {
        "id": "W0002",
        "nama": "Contoh Nama Istri", "jenisKelamin": "PEREMPUAN",
        "tempatLahir": "Bandung", "tanggalLahir": "1992-05-02", "agama": "ISLAM",
        "statusPerkawinan": "KAWIN", "pendidikan": "SMA",
        "pekerjaan": "Ibu Rumah Tangga", "golonganDarah": "A",
        "statusHubunganKeluarga": "ISTRI", "kewarganegaraan": "WNI",
        "jabatan": "WARGA",
        "jalan": "Jl. Contoh No. 1", "rt": "001", "rw": "019", "desa": "Sukamaju",
        "kecamatan": "Cibiru", "kabupaten": "Bandung", "provinsi": "Jawa Barat",
        "kodePos": "40615",
    },
]

PETUNJUK = [
    "PETUNJUK PENGISIAN",
    "",
    "1. Jangan ubah JUDUL kolom di sheet 'Data Penduduk'. Urutannya boleh digeser.",
    "2. Dua baris contoh (huruf miring abu-abu) HAPUS sebelum dipakai.",
    "3. NIK dan No. KK TIDAK didata. Jangan menambahkan kolomnya sendiri —",
    "   sistem tidak menyimpannya, dan desa tidak mengizinkannya.",
    "",
    "   KODE WARGA — kolom paling penting:",
    "   - Wajib diisi, tidak boleh kosong, tidak boleh sama dengan warga lain.",
    "   - Sekali diberikan, JANGAN PERNAH diubah. Kode inilah yang dipakai",
    "     sistem untuk mengenali orang yang sama pada pendataan berikutnya.",
    "   - Warga baru mendapat kode berikutnya yang belum terpakai; kode milik",
    "     warga yang meninggal atau pindah JANGAN dipakai ulang.",
    "4. Tanggal Lahir: format yyyy-mm-dd (contoh 1990-01-01).",
    "5. Kolom berikut WAJIB pilih dari dropdown (klik sel, muncul panah di kanan):",
    "   Jenis Kelamin, Agama, Status Perkawinan, Pendidikan Terakhir,",
    "   Gol. Darah, Status dalam Keluarga.",
    "",
    "   JABATAN — isi WARGA untuk hampir semua orang.",
    "   - DUKUH untuk Pak Dukuh, RW untuk Ketua RW, RT untuk Ketua RT.",
    "   - Satu jabatan satu orang: jangan ada dua RT di RT yang sama.",
    "   - Dipakai HANYA saat jabatannya masih kosong di aplikasi. Setelah",
    "     jabatannya terisi, pergantian dilakukan lewat aplikasi (butuh",
    "     persetujuan), dan kolom ini tidak lagi berpengaruh.",
    "",
    "6. Kirim file LENGKAP berisi SELURUH warga, bukan tambahannya saja:",
    "   impor menimpa seluruh data lama.",
    "7. Simpan sebagai .xlsx, kirim ke pengelola data untuk diimpor ke sistem.",
]


def main() -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = NAMA_SHEET

    tipis = Side(style="thin", color="B7B7B7")
    bingkai = Border(left=tipis, right=tipis, top=tipis, bottom=tipis)

    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(KOLOM))
    judul = ws.cell(row=1, column=1, value="FORMULIR PENDATAAN PENDUDUK — DESA SUKAMAJU")
    judul.fill = PatternFill("solid", fgColor="1F4E78")
    judul.font = Font(color="FFFFFF", bold=True, size=13)
    judul.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 24

    for i, (_, label, lebar) in enumerate(KOLOM, start=1):
        c = ws.cell(row=BARIS_HEADER, column=i, value=label)
        c.fill = PatternFill("solid", fgColor="D9E1F2")
        c.font = Font(bold=True)
        c.border = bingkai
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        ws.column_dimensions[get_column_letter(i)].width = lebar
    ws.row_dimensions[BARIS_HEADER].height = 30
    ws.freeze_panes = f"A{BARIS_HEADER + 1}"

    baris_1 = BARIS_HEADER + 1
    for r, contoh in enumerate(CONTOH, start=baris_1):
        for i, (field, _, _) in enumerate(KOLOM, start=1):
            c = ws.cell(row=r, column=i, value=contoh[field])
            c.font = Font(italic=True, color="808080")
            c.border = bingkai

    baris_akhir = baris_1 + len(CONTOH) + BARIS_KOSONG
    for r in range(baris_1 + len(CONTOH), baris_akhir + 1):
        for i in range(1, len(KOLOM) + 1):
            ws.cell(row=r, column=i).border = bingkai

    for i, (field, _, _) in enumerate(KOLOM, start=1):
        if field not in PILIHAN:
            continue
        opsi = PILIHAN[field]
        huruf = get_column_letter(i)
        dv = DataValidation(
            type="list", formula1=f'"{",".join(opsi)}"', allow_blank=True,
            showErrorMessage=True, errorTitle="Nilai tidak valid",
            error="Pilih salah satu: " + ", ".join(opsi),
        )
        ws.add_data_validation(dv)
        dv.add(f"{huruf}{baris_1}:{huruf}{baris_akhir}")

    ws2 = wb.create_sheet("Petunjuk")
    for i, teks in enumerate(PETUNJUK, start=1):
        c = ws2.cell(row=i, column=1, value=teks)
        if i == 1:
            c.font = Font(bold=True, size=13)
    ws2.column_dimensions["A"].width = 90

    wb.save(KELUARAN)
    print(f"OK: {KELUARAN} ({len(KOLOM)} kolom, {BARIS_KOSONG} baris kosong siap isi)")


if __name__ == "__main__":
    main()
