"""Isi tabel penduduk dari file Excel hasil pendataan pengurus. Kolomnya
dibaca lewat app/schemas/penduduk.py, disimpan lewat app/data/db.py (sudah
tervalidasi self-check-nya sendiri).

Dipakai untuk **mengisi data pertama kali**. Sejak Tahap 3b, sumber kebenaran
data warga adalah aplikasi — pengurus mengubah dan menambah lewat layar, bukan
lewat file ini.

Setiap impor MENIMPA seluruh tabel penduduk, jadi skrip ini **menolak jalan
kalau database sudah berisi** kecuali diberi `--timpa-semua`.

Tidak ada dedup: tanpa NIK tidak ada kunci yang bisa dipercaya untuk mengenali
orang yang sama antar-impor, dan kandidat penggantinya (nama + tanggal lahir +
alamat) gagal persis pada kasus yang paling mungkin terjadi di satu padukuhan:
dua orang senama.

Kolom dicocokkan lewat **nama header**, bukan urutan. Pengurus boleh
menggeser, jadi mengandalkan urutan berarti data masuk ke kolom yang salah
tanpa ada yang menyadarinya.

Backend boleh tetap menyala saat skrip ini jalan — SQLite aman ditulis
sementara proses lain membacanya, dan sejak Tahap 3a tidak ada cache: hasilnya
langsung kelihatan di API tanpa restart.

Pakai:
    .venv/bin/pip install openpyxl   # sekali, alat ini saja yang butuh
    .venv/bin/python -m app.data.impor_excel ../docs/data-penduduk.xlsx
"""

import sys

from openpyxl import load_workbook

from app.core.config import settings
from app.data import db
from app.schemas.penduduk import Alamat, Penduduk

NAMA_SHEET = "Data Penduduk"
BARIS_HEADER = 2  # baris 1 = judul

# Satu-satunya definisi kolom: (field Pydantic, label di Excel, lebar kolom).
# Pembangkit file Excel di `backend/tools/` membaca daftar ini juga, supaya
# bentuk file dan pembacanya tidak mungkin melenceng sendiri-sendiri.
KOLOM: list[tuple[str, str, int]] = [
    ("id", "Kode Warga", 14),
    ("nama", "Nama Lengkap", 24),
    ("jenisKelamin", "Jenis Kelamin", 14),
    ("tempatLahir", "Tempat Lahir", 16),
    ("tanggalLahir", "Tanggal Lahir (yyyy-mm-dd)", 20),
    ("agama", "Agama", 12),
    ("statusPerkawinan", "Status Perkawinan", 16),
    ("pendidikan", "Pendidikan Terakhir", 14),
    ("pekerjaan", "Pekerjaan", 20),
    ("golonganDarah", "Gol. Darah", 10),
    ("statusHubunganKeluarga", "Status dalam KK", 18),
    ("kewarganegaraan", "Kewarganegaraan", 14),
    ("jabatan", "Jabatan", 12),
    ("jalan", "Alamat Jalan", 26),
    ("rt", "RT", 6),
    ("rw", "RW", 6),
    ("desa", "Desa/Kelurahan", 16),
    ("kecamatan", "Kecamatan", 14),
    ("kabupaten", "Kabupaten", 14),
    ("provinsi", "Provinsi", 14),
    ("kodePos", "Kode Pos", 10),
]

# Kolom yang isinya terkunci ke sedikit pilihan. Dipakai pembangkit template
# untuk memasang dropdown; salah ketik ditangkap Pydantic saat impor.
PILIHAN: dict[str, list[str]] = {
    "jenisKelamin": ["LAKI_LAKI", "PEREMPUAN"],
    "agama": ["ISLAM", "KRISTEN", "KATOLIK", "HINDU", "BUDDHA", "KONGHUCU", "LAINNYA"],
    "statusPerkawinan": ["BELUM_KAWIN", "KAWIN", "CERAI_HIDUP", "CERAI_MATI"],
    "pendidikan": ["TIDAK_SEKOLAH", "SD", "SMP", "SMA", "D3", "S1", "S2", "S3"],
    "golonganDarah": ["A", "B", "AB", "O", "TIDAK_TAHU"],
    "statusHubunganKeluarga": [
        "KEPALA_KELUARGA", "ISTRI", "ANAK", "FAMILI_LAIN", "LAINNYA",
    ],
    "jabatan": ["WARGA", "DUKUH", "RW", "RT"],
}

_ALAMAT = {
    "jalan", "rt", "rw", "desa", "kecamatan", "kabupaten", "provinsi", "kodePos",
}


def _samakan(teks: object) -> str:
    """Header dibandingkan tanpa peduli huruf besar-kecil & spasi berlebih."""
    return " ".join(str(teks or "").split()).lower()


def petakan_kolom(baris_header: tuple) -> dict[str, int]:
    """Nama field -> indeks kolom, dicocokkan dari label di baris header."""
    ada = {_samakan(v): i for i, v in enumerate(baris_header) if v is not None}
    peta, hilang = {}, []
    for field, label, _ in KOLOM:
        i = ada.get(_samakan(label))
        if i is None:
            hilang.append(label)
        else:
            peta[field] = i
    if hilang:
        sys.exit(
            "Kolom berikut tidak ketemu di baris header Excel:\n  "
            + "\n  ".join(hilang)
            + "\n\nJudul kolom tidak boleh diubah — urutannya boleh."
        )
    return peta


def baris_ke_penduduk(nilai: dict[str, str]) -> Penduduk:
    inti = {k: v for k, v in nilai.items() if k not in _ALAMAT}
    alamat = {k: v for k, v in nilai.items() if k in _ALAMAT}
    # `id` = kolom "Kode Warga", kunci yang dijaga pengurus. Bukan dibangkitkan
    # acak: jabatan pengurus menunjuk ke warga tertentu, dan impor menimpa
    # seluruh tabel — id acak akan memutus tautan itu tiap kali impor.
    return Penduduk(alamat=Alamat(**alamat), **inti)


def baca_xlsx(path: str) -> list[Penduduk]:
    ws = load_workbook(path, data_only=True)[NAMA_SHEET]
    baris = list(ws.iter_rows(min_row=BARIS_HEADER, values_only=True))
    peta = petakan_kolom(baris[0])

    daftar: list[Penduduk] = []
    kosong: list[int] = []
    baris_ke_nomor: dict[str, list[int]] = {}
    # +1 karena `baris` dimulai dari baris header, dan Excel menghitung dari 1.
    for nomor, r in enumerate(baris[1:], start=BARIS_HEADER + 1):
        if not r[peta["nama"]] and not r[peta["id"]]:
            continue  # dua-duanya kosong = baris belum diisi, lewati
        nilai = {
            field: ("" if r[i] is None else str(r[i]).strip())
            for field, i in peta.items()
        }
        baris_ke_nomor.setdefault(nilai["id"], []).append(nomor)
        if not nilai["id"]:
            kosong.append(nomor)
            continue
        daftar.append(baris_ke_penduduk(nilai))

    if kosong:
        sys.exit(
            "Kolom 'Kode Warga' kosong di baris: "
            + ", ".join(str(b) for b in kosong)
            + "\n\nKode Warga adalah identitas warga di sistem — tanpa itu barisnya "
            "tidak bisa dipakai. Isi dulu, lalu jalankan ulang."
        )

    _periksa_jabatan(daftar, baris_ke_nomor)

    ganda = {k: b for k, b in _cari_ganda(daftar, baris_ke_nomor).items()}
    if ganda:
        sys.exit(
            "Kode Warga berikut dipakai lebih dari satu baris:\n  "
            + "\n  ".join(f"{kode} (baris {', '.join(map(str, b))})" for kode, b in ganda.items())
            + "\n\nSatu kode = satu orang. Dua warga bertukar kode berarti dua orang "
            "bertukar jabatan tanpa ada yang menyadarinya. Betulkan dulu."
        )
    return daftar


def _periksa_jabatan(
    daftar: list[Penduduk], nomor_baris: dict[str, list[int]]
) -> None:
    """Satu jabatan satu orang, sudah di file Excel-nya.

    Dua orang bertanda `RT` di RT yang sama berarti file itu sendiri tidak tahu
    siapa ketuanya — dan aplikasi tidak boleh menebak. Diperiksa sebelum satu
    baris pun ditulis.
    """
    pemegang: dict[str, list[str]] = {}
    for p in daftar:
        if p.jabatan == "WARGA":
            continue
        if p.jabatan == "DUKUH":
            jabatan = "Dukuh"
        elif p.jabatan == "RW":
            jabatan = f"Ketua RW {p.alamat.rw}"
        else:
            jabatan = f"Ketua RT {p.alamat.rt} (RW {p.alamat.rw})"
        pemegang.setdefault(jabatan, []).append(p.nama)

    bentrok = {k: v for k, v in pemegang.items() if len(v) > 1}
    if bentrok:
        sys.exit(
            "Satu jabatan ditandai untuk lebih dari satu orang:\n  "
            + "\n  ".join(f"{k}: {', '.join(v)}" for k, v in bentrok.items())
            + "\n\nSatu jabatan dipegang satu orang. Betulkan kolom Jabatan dulu."
        )


def _cari_ganda(
    daftar: list[Penduduk], nomor_baris: dict[str, list[int]]
) -> dict[str, list[int]]:
    """Kode yang muncul lebih dari sekali, beserta nomor barisnya."""
    return {p.id: nomor_baris[p.id] for p in daftar if len(nomor_baris[p.id]) > 1}


def main(path_xlsx: str, timpa: bool = False) -> None:
    daftar = baca_xlsx(path_xlsx)
    if not daftar:
        sys.exit("Tidak ada baris terisi — cek lagi apakah baris contoh sudah dihapus.")

    conn = db.buka(settings.DATABASE_FILE)

    # Sejak Tahap 3b aplikasi adalah sumber kebenaran data warga, bukan Excel.
    # Tanpa kunci ini, satu perintah impor yang dijalankan karena kebiasaan
    # menghapus seluruh hasil pendataan yang dikerjakan pengurus di aplikasi.
    if not db.kosong(conn) and not timpa:
        jumlah = len(db.muat(conn))
        conn.close()
        sys.exit(
            f"Database sudah berisi {jumlah} warga, dan impor MENIMPA semuanya.\n\n"
            "Sejak pengurus bisa mengubah data lewat aplikasi, isi database bisa "
            "lebih baru daripada file Excel — termasuk warga baru dan koreksi "
            "yang tidak ada di file ini.\n\n"
            "Kalau memang mau membuang seluruh isi database dan menggantinya "
            "dengan file ini, ulangi dengan:\n"
            f"  .venv/bin/python -m app.data.impor_excel {path_xlsx} --timpa-semua\n\n"
            "Salin dulu file databasenya sebelum itu."
        )

    lama = db.kosongkan(conn)
    masuk = db.simpan(conn, daftar)
    conn.close()
    if lama:
        print(f"PERHATIAN: {lama} baris lama dihapus dan diganti seluruhnya.")
    print(f"OK: {masuk} warga masuk ke {settings.DATABASE_FILE}")
    print("Langsung kepakai — backend query database tiap request, tidak ada cache.")


if __name__ == "__main__":
    arg = [a for a in sys.argv[1:] if a != "--timpa-semua"]
    if len(arg) != 1:
        sys.exit(
            "Pakai: python -m app.data.impor_excel <path-ke-file.xlsx> [--timpa-semua]"
        )
    main(arg[0], timpa="--timpa-semua" in sys.argv)
